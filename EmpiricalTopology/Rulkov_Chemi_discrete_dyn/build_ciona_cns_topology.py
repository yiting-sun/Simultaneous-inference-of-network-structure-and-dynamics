"""
Build the 177-node ciona CNS binary adjacency used by the Rulkov discrete-map
generation and training pipelines, with self-loops removed.

The eLife Figure-16 workbook encodes row = presynaptic cell, column =
postsynaptic partner. We restrict to the 177 "CNS neurons only" subset and
convert to A[i, j] = 1 meaning a directed edge j -> i (the convention used
throughout the pipeline).

Two differences from the parent project:
  (1) We REMOVE all self-loops (A[i, i] = 0) so the Rulkov message passing
      does not see a node sending signal to itself.
  (2) We report the number of self-loops stripped in the metadata as a
      sanity check.
"""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
RAW_MATRIX = DATA_DIR / "elife-16962-fig16-data1-v1.xlsx"

OUT_NPY = DATA_DIR / "ciona_cns_177_binary_adj.npy"
OUT_CSV = DATA_DIR / "ciona_cns_177_binary_adj.csv"
OUT_LABELS = DATA_DIR / "ciona_cns_177_labels.txt"
OUT_META = DATA_DIR / "ciona_cns_177_metadata.json"


# Figure 16 source-data matrix is a 205 x 215 directed connectivity matrix:
# rows are presynaptic cells and columns are postsynaptic partners.
# The paper reports 177 "CNS neurons only" nodes. We recover that subset by
# removing the non-CNS presynaptic entries from the row labels.
NON_CNS_PRESYNAPTIC = {
    "pns1", "pns10", "pns11", "pns12", "pns13", "pns3", "pns4", "pns5",
    "pns6", "pns7", "pns9",
    "ATEN1", "ATEN2", "ATEN3", "ATEN4",
    "pna", "pnb", "pnc", "pnf", "pnh", "pnu", "pnw", "pnx",
    "BTN1", "BTN2", "BTN3", "BTN4",
    "lens6",
}

# The raw workbook has a few row/column naming inconsistencies. These remaps
# align the selected 177 row labels to the corresponding postsynaptic columns.
POSTSYNAPTIC_COLUMN_REMAP = {
    19: 5,
    40: 35,
    44: 57,
    "coronet10": "coronet14",
    "coronet5": "coronet2",
}


def load_raw_labels(ws):
    row_labels = [
        ws.cell(r, 1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value not in (None, "Total")
    ]
    col_labels = [
        ws.cell(1, c).value
        for c in range(2, ws.max_column + 1)
        if ws.cell(1, c).value not in (None, "Total")
    ]
    return row_labels, col_labels


def build_binary_adjacency():
    if not RAW_MATRIX.exists():
        raise FileNotFoundError(
            f"Missing raw matrix file: {RAW_MATRIX}. Download from the eLife article first."
        )

    workbook = openpyxl.load_workbook(RAW_MATRIX, data_only=True)
    ws = workbook["Sheet1"]
    row_labels, col_labels = load_raw_labels(ws)

    labels = [label for label in row_labels if label not in NON_CNS_PRESYNAPTIC]
    if len(labels) != 177:
        raise ValueError(f"Expected 177 CNS labels, got {len(labels)}")

    row_index = {label: idx + 2 for idx, label in enumerate(row_labels)}
    col_index = {label: idx + 2 for idx, label in enumerate(col_labels)}

    mapped_columns = [POSTSYNAPTIC_COLUMN_REMAP.get(label, label) for label in labels]
    if len(set(mapped_columns)) != len(labels):
        raise ValueError("Postsynaptic column remap is not one-to-one")

    missing_cols = [label for label in mapped_columns if label not in col_index]
    if missing_cols:
        raise KeyError(f"Missing mapped postsynaptic columns: {missing_cols}")

    adjacency = np.zeros((len(labels), len(labels)), dtype=np.uint8)

    # Workbook encodes row -> column. Convert to A[i, j] where j -> i.
    for pre_j, pre_label in enumerate(labels):
        row_id = row_index[pre_label]
        for post_i, post_label in enumerate(labels):
            col_id = col_index[POSTSYNAPTIC_COLUMN_REMAP.get(post_label, post_label)]
            value = ws.cell(row_id, col_id).value
            adjacency[post_i, pre_j] = 1 if isinstance(value, (int, float)) and value > 0 else 0

    return adjacency, labels


def remove_self_loops(adjacency):
    """Zero out the diagonal and return (clean_adj, num_self_loops_removed)."""
    num_self_loops = int(np.trace(adjacency))
    clean = adjacency.copy()
    np.fill_diagonal(clean, 0)
    return clean, num_self_loops


def verify_no_self_loops(adjacency):
    """Assert the adjacency has zero self-loops (used both after cleaning and by
    downstream pipeline entry points as a sanity gate)."""
    diag = int(np.trace(adjacency))
    if diag != 0:
        raise ValueError(
            f"Adjacency still has {diag} self-loop(s); refuse to save / use."
        )


def save_outputs(adjacency, labels, num_self_loops_removed):
    label_strings = [str(label) for label in labels]

    verify_no_self_loops(adjacency)

    np.save(OUT_NPY, adjacency)
    pd.DataFrame(adjacency, index=label_strings, columns=label_strings).to_csv(OUT_CSV)
    OUT_LABELS.write_text("\n".join(label_strings) + "\n", encoding="utf-8")

    metadata = {
        "source_article": "https://elifesciences.org/articles/16962",
        "source_matrix_file": RAW_MATRIX.name,
        "matrix_semantics": "A[i,j] = 1 means a directed edge j -> i",
        "raw_matrix_semantics": "Workbook rows are presynaptic cells and columns are postsynaptic partners",
        "node_count": int(adjacency.shape[0]),
        "edge_count": int(adjacency.sum()),
        "density": float(adjacency.mean()),
        "self_loops_in_saved_adjacency": int(np.trace(adjacency)),
        "self_loops_removed_from_raw": int(num_self_loops_removed),
        "excluded_non_cns_presynaptic_labels": [str(x) for x in sorted(NON_CNS_PRESYNAPTIC, key=str)],
        "postsynaptic_column_remap": {str(k): str(v) for k, v in POSTSYNAPTIC_COLUMN_REMAP.items()},
        "output_files": {
            "npy": OUT_NPY.name,
            "csv": OUT_CSV.name,
            "labels": OUT_LABELS.name,
        },
    }
    OUT_META.write_text(json.dumps(metadata, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main():
    raw_adjacency, labels = build_binary_adjacency()
    adjacency, num_self_loops_removed = remove_self_loops(raw_adjacency)
    save_outputs(adjacency, labels, num_self_loops_removed)

    print(f"Saved {adjacency.shape[0]} x {adjacency.shape[1]} binary adjacency to {OUT_NPY}")
    print(f"Edges (after self-loop removal): {int(adjacency.sum())}")
    print(f"Self-loops stripped: {num_self_loops_removed}")
    print(f"Density: {adjacency.mean():.6f}")


if __name__ == "__main__":
    main()
